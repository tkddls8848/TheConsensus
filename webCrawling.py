from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import Workbook # 엑셀 저장 모듈
import time
import datetime

# 현재 시간
now = datetime.datetime.now()
nowDate = "{}년{}월{}일".format(now.year, now.month, now.day)

# 웹 드라이버
driver = webdriver.Chrome("./chromedriver")
# 엑셀 워크북 로드
write_workbook = Workbook()
write_cell = write_workbook.active

# 받은 문자 리스트
result_list = []

# 여러 페이지 받기
for i in range(1,3): # i는 페이지 수
    driver.get("https://gall.dcinside.com/mgallery/board/lists/?id=jusik&page=" + str(i))
    time.sleep(0.1) # 페이지 꼬임 방지
    
    # 받은 코드 html로 변환
    soup = BeautifulSoup(driver.page_source, "html.parser")

    # html 코드 중 게시글의 제목만 받기
    for j in soup.select("#container > section.left_content > article:nth-child(3) > div.gall_listwrap.list > table > tbody > tr"):
        print(j.find("td", class_="gall_tit ub-word").text)
        result_list.append(j.find("td", class_="gall_tit ub-word").text[1:].split('\n')[0])

    for k in range(1, len(result_list) + 1):
        write_cell.cell(k, 1, str(i) + " 페이지")
        write_cell.cell(k, 2, result_list[k - 1])
    write_workbook.save("실투갤게시판_" + str(i) + "_PAGE_" + nowDate + "제목리스트.xls")
    result_list.clear()

print(result_list)

#웹 드라이버 종료
driver.close()
#엑셀 워크북 종료
write_workbook.close()

