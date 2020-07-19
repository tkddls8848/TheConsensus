from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import Workbook # 엑셀 저장 모듈
import time
import datetime
from konlpy.tag import Kkma
import collections

# 현재 시간
now = datetime.datetime.now()
nowDate = "{}년{}월{}일".format(now.year, now.month, now.day)

# 웹 드라이버
driver = webdriver.Chrome("../chromedriver")
# 엑셀 워크북 로드
write_workbook = Workbook()
write_cell = write_workbook.active
# Konlpy 모듈 로드
kkma = Kkma()

# 받은 문자 페이지별 리스트
result_list = []
# 명사 분류 리스트의 명사별 카운트
count_list = []

# 여러 페이지 받기
for i in range(1,3): # i는 페이지 수
    driver.get("https://gall.dcinside.com/mgallery/board/lists/?id=jusik&page=" + str(i))
    time.sleep(0.1) # 페이지 꼬임 방지
    
    # 받은 코드 html로 변환
    soup = BeautifulSoup(driver.page_source, "html.parser")
    # html 코드 중 게시글의 제목만 받기
    for j in soup.select(
        "#container > section.left_content > article:nth-of-type(2) > div.gall_listwrap.list > table > tbody > tr"):
        print(j.find("td", class_="gall_tit ub-word").text)
        add_text = j.find("td", class_="gall_tit ub-word").text[1:].split('\n')[0]
        result_list.append(add_text)

    # 게시글의 제목 엑셀 저장
    for k in range(1, len(result_list) + 1):
        write_cell.cell(k, 1, str(i) + " 페이지")
        write_cell.cell(k, 2, result_list[k - 1])
    write_workbook.save("JusikGall_" + str(i) + "_PAGE_" + nowDate + "_List.xlsx")
    
    #리스트를 뽑아서 명사분해하여 다시 리스트
    for l in result_list:
        count_list = count_list + kkma.nouns(l)
        
    # 제목 추출 결과 많이 나온 5위까지의 명사
    print(collections.Counter(count_list).most_common(5))
    
    # result 초기화
    result_list.clear()

# 웹 드라이버 종료
driver.close()

# 엑셀 워크북 종료
write_workbook.close()

