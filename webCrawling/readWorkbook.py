from openpyxl import load_workbook
from konlpy.tag import Kkma
import datetime
import collections
import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm

# 현재 시간
now = datetime.datetime.now()
nowDate = "{}년{}월{}일".format(now.year, now.month, now.day)

# Kkma로드
kkma = Kkma()

# 엑셀 파일별 로드
for a in range(1, 3):

    # 엑셀 로드
#   read_workbook = load_workbook('./JusikGall_' + str(a) + '_PAGE_' + nowDate + '_List.xlsx')
    read_workbook = load_workbook('./JusikGall_1_PAGE_2020년7월21일_List.xlsx')
    read_cell = read_workbook.active

    # 결과 리스트
    result_list = []

    # 리스트 분석
    for i in range(6, 15) :
        print(read_cell.cell(i, 2).value)
        # 개별 제목 명사 분해
        for j in kkma.nouns(read_cell.cell(i, 2).value):
            # 조사 삭제
            if len(kkma.nouns(read_cell.cell(i, 2).value)) != 1:
                result_list.append(j)
        print(kkma.nouns(read_cell.cell(i, 2).value))

# 결과 count
print(collections.Counter(result_list))

# 결과 카운터 딕셔너리를 키 값 리스트로 변환
result = collections.Counter(result_list)
result_key = list(result.keys())
result_value = list(result.values())

# 한글 폰트 로드
font_location = 'C:\WINDOWS\FONTS\MALGUNSL.TTF'
font_name = fm.FontProperties(fname=font_location).get_name()
mpl.rc('font', family=font_name)

# 데이터 그래프 표현
label = result_key
index = np.arange(len(label))
plt.bar(index, result_value)

plt.title('title')
plt.xlabel('x축')
plt.ylabel('y축')
plt.xticks(index, label)

plt.show()
