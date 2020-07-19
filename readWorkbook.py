from openpyxl import load_workbook
from konlpy.tag import Kkma
import datetime
import collections

# 현재 시간
now = datetime.datetime.now()
nowDate = "{}년{}월{}일".format(now.year, now.month, now.day)

# Kkma로드
kkma = Kkma()

# 엑셀 로드
read_workbook = load_workbook('./JusikGall_1_PAGE_' + nowDate + '_List.xlsx')
read_cell = read_workbook.active

# 결과 리스트
result_list = []

# 리스트 분석
for i in range(6, 20) :
    print(read_cell.cell(i, 2).value)
    # 개별 제목 명사 분해
    for j in kkma.nouns(read_cell.cell(i, 2).value) :
        result_list.append(j)
    print(kkma.nouns(read_cell.cell(i, 2).value))

print(collections.Counter(result_list))